"Reviews lists."

import tempfile

import flask
import openpyxl
import openpyxl.styles

import anubis.call
import anubis.user
import anubis.proposal
import anubis.review

from . import constants
from . import utils

blueprint = flask.Blueprint('reviews', __name__)

@blueprint.route('/call/<cid>')
@utils.login_required
def call(cid):
    "List all reviews for a call."
    call = anubis.call.get_call(cid)
    if call is None:
        return utils.error('No such call.', flask.url_for('home'))
    if not anubis.call.allow_view(call):
        return utils.error('You may not view the call.', flask.url_for('home'))
    if not anubis.call.allow_view_reviews(call):
        return utils.error('You may not view the reviews of the call.',
                           flask.url_for('call.display',cid=call['identifier']))

    proposals = utils.get_docs_view('proposals', 'call', call['identifier'])
    for proposal in proposals:
        proposal['allow_create_review'] = anubis.review.allow_create(proposal)
    reviews = utils.get_docs_view('reviews', 'call', call['identifier'])
    # For ordinary reviewer, list only finalized reviews.
    if not (flask.g.am_admin or anubis.call.am_chair(call)):
        finalized = True
        reviews = [r for r in reviews
                   if r['reviewer'] != flask.g.current_user['username'] and 
                   r.get('finalized')]
    else:
        finalized = False
    reviews_lookup = {f"{r['proposal']} {r['reviewer']}":r for r in reviews}
    review_bannerfields = [f for f in call['review'] if f.get('banner')]
    result = flask.render_template('reviews/call.html',
                                 call=call,
                                 proposals=proposals,
                                 reviews_lookup=reviews_lookup,
                                 finalized=finalized,
                                 review_bannerfields=review_bannerfields)
    return result

@blueprint.route('/call/<cid>.xlsx')
@utils.login_required
def call_xlsx(cid):
    "Produce an XLSX file of all reviews for a call."
    call = anubis.call.get_call(cid)
    if call is None:
        return utils.error('No such call.', flask.url_for('home'))
    if not anubis.call.allow_view(call):
        return utils.error('You may not view the call.', flask.url_for('home'))
    if not anubis.call.allow_view_reviews(call):
        return utils.error('You may not view the reviews of the call.',
                           flask.url_for('call.display',cid=call['identifier']))

    proposals = utils.get_docs_view('proposals', 'call', call['identifier'])
    reviews = utils.get_docs_view('reviews', 'call', call['identifier'])
    # For ordinary reviewer, list only finalized reviews.
    if not (flask.g.am_admin or anubis.call.am_chair(call)):
        reviews = [r for r in reviews
                   if r['reviewer'] != flask.g.current_user['username'] and 
                   r.get('finalized')]
    reviews_lookup = {f"{r['proposal']} {r['reviewer']}":r for r in reviews}
    content = get_xlsx(call, proposals, reviews_lookup)
    response = flask.make_response(content)
    response.headers.set('Content-Type', constants.XLSX_MIMETYPE)
    response.headers.set('Content-Disposition', 'attachment', 
                         filename=f"{cid}_reviews.xlsx")
    return response

@blueprint.route('/call/<cid>/reviewer/<username>')
@utils.login_required
def call_reviewer(cid, username):
    "List all reviews in the call by the reviewer (user)."
    call = anubis.call.get_call(cid)
    if call is None:
        return utils.error('No such call.', flask.url_for('home'))
    user = anubis.user.get_user(username=username)
    if user is None:
        return utils.error('No such user.', flask.url_for('home'))
    if user['username'] not in call['reviewers']:
        return utils.error('The user is not a reviewer in the call.',
                           flask.url_for('home'))
    if not (user['username'] == flask.g.current_user['username']  or
            anubis.call.allow_view_reviews(call)):
        return utils.error("You may not view the user's reviews.",
                           flask.url_for('call.display',cid=call['identifier']))

    proposals = utils.get_docs_view('proposals', 'call', call['identifier'])
    reviews = utils.get_docs_view('reviews', 'call_reviewer',
                                  [call['identifier'], user['username']])
    reviews_lookup = {r['proposal']:r for r in reviews}
    bannerfields = [f for f in call['review'] if f.get('banner')]
    return flask.render_template('reviews/call_reviewer.html', 
                                 call=call,
                                 proposals=proposals,
                                 user=user,
                                 reviews_lookup=reviews_lookup,
                                 bannerfields=bannerfields)

@blueprint.route('/call/<cid>/reviewer/<username>.xlsx')
@utils.login_required
def call_reviewer_xlsx(cid, username):
    "Produce an XLSX file of all reviews in the call by the reviewer (user)."
    call = anubis.call.get_call(cid)
    if call is None:
        return utils.error('No such call.', flask.url_for('home'))
    user = anubis.user.get_user(username=username)
    if user is None:
        return utils.error('No such user.', flask.url_for('home'))
    if user['username'] not in call['reviewers']:
        return utils.error('The user is not a reviewer in the call.',
                           flask.url_for('home'))
    if not (user['username'] == flask.g.current_user['username']  or
            anubis.call.allow_view_reviews(call)):
        return utils.error("You may not view the user's reviews.",
                           flask.url_for('call.display',cid=call['identifier']))

    proposals = utils.get_docs_view('proposals', 'call', call['identifier'])
    reviews = utils.get_docs_view('reviews', 'call_reviewer',
                                  [call['identifier'], user['username']])
    reviews_lookup = {f"{r['proposal']} {username}":r for r in reviews}
    content = get_xlsx(call, proposals, reviews_lookup)
    response = flask.make_response(content)
    response.headers.set('Content-Type', constants.XLSX_MIMETYPE)
    response.headers.set('Content-Disposition', 'attachment', 
                         filename=f"{cid}_{username}_reviews.xlsx")
    return response

@blueprint.route('/proposal/<pid>')
@utils.login_required
def proposal(pid):
    "List all reviewers and reviews for a proposal."
    proposal = anubis.proposal.get_proposal(pid)
    if proposal is None:
        return utils.error('No such proposal.', flask.url_for('home'))

    call = anubis.call.get_call(proposal['call'])
    if not anubis.call.allow_view_reviews(call):
        return utils.error('You may not view the reviews of the call.',
                           flask.url_for('call.display',cid=call['identifier']))

    reviews = utils.get_docs_view('reviews', 'proposal', proposal['identifier'])
    if not (flask.g.am_admin or anubis.call.am_chair(call)):
        finalized = True
        reviews = [r for r in reviews
                   if r['reviewer'] != flask.g.current_user['username'] and 
                   r.get('finalized')]
    else:
        finalized = False
    allow_create_review = anubis.review.allow_create(proposal)
    reviews_lookup = {r['reviewer']:r for r in reviews}
    bannerfields = [f for f in call['review'] if f.get('banner')]
    return flask.render_template('reviews/proposal.html',
                                 proposal=proposal,
                                 call=call,
                                 allow_create_review=allow_create_review,
                                 reviewers=call['reviewers'],
                                 reviews_lookup=reviews_lookup,
                                 finalized=finalized,
                                 bannerfields=bannerfields)

@blueprint.route('/proposal/<pid>.xlsx')
@utils.login_required
def proposal_xlsx(pid):
    "Produce an XLSX file of all reviewers and reviews for a proposal."
    proposal = anubis.proposal.get_proposal(pid)
    if proposal is None:
        return utils.error('No such proposal.', flask.url_for('home'))
    if not anubis.proposal.allow_view(proposal):
        return utils.error('You may not view the proposal.',
                           flask.url_for('call.display',cid=call['identifier']))
    call = anubis.call.get_call(proposal['call'])
    if not anubis.call.allow_view_reviews(call):
        return utils.error('You may not view the reviews of the call.',
                           flask.url_for('call.display',cid=call['identifier']))

    reviews = utils.get_docs_view('reviews', 'proposal', proposal['identifier'])
    if not (flask.g.am_admin or anubis.call.am_chair(call)):
        reviews = [r for r in reviews
                   if r['reviewer'] != flask.g.current_user['username'] and 
                   r.get('finalized')]
    reviews_lookup = {f"{pid} {r['reviewer']}":r for r in reviews}
    content = get_xlsx(call, [proposal], reviews_lookup)
    response = flask.make_response(content)
    response.headers.set('Content-Type', constants.XLSX_MIMETYPE)
    response.headers.set('Content-Disposition', 'attachment', 
                         filename=f"{pid}_reviews.xlsx")
    return response

@blueprint.route('/reviewer/<username>')
@utils.login_required
def reviewer(username):
    """List all reviews by the given reviewer (user).
    If the user is reviewer in only one call, redirect to that page.
    """
    user = anubis.user.get_user(username=username)
    if user is None:
        return utils.error('No such user.', flask.url_for('home'))
    if not anubis.user.allow_view(user):
        return utils.error("You may not view the user's reviews.",
                           flask.url_for('call.display',cid=call['identifier']))

    reviewer_calls = [anubis.call.get_call(r.value)
                      for r in flask.g.db.view('calls', 'reviewer', 
                                               key=user['username'],
                                               reduce=False)]
    # Reviews in only one call; redirect to its reviews page for the reviewer.
    if len(reviewer_calls) == 1:
        return flask.redirect(flask.url_for('reviews.call_reviewer',
                                            cid=reviewer_calls[0]['identifier'],
                                            username=username))

    reviews = utils.get_docs_view('reviews', 'reviewer', user['username'])
    return flask.render_template('reviews/reviewer.html',
                                 user=user,
                                 reviewer_calls=reviewer_calls,
                                 reviews=reviews)

def get_xlsx(call, proposals, reviews_lookup):
    "Return the content for the XLSX file for the list of reviews."
    from openpyxl.utils.cell import get_column_letter
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Reviews in call {call['identifier']}"
    row = ['Proposal', 'Proposal title', 'Submitter', 
           'Review', 'Finalized', 'Reviewer']
    ws.column_dimensions[get_column_letter(2)].width = 30
    review_column = get_column_letter(4)
    ws.column_dimensions[review_column].width = 30
    for field in call['review']:
        row.append(field['identifier'])
        if field['type'] == constants.LINE:
            ws.column_dimensions[get_column_letter(len(row))].width = 40
        elif field['type'] == constants.TEXT:
            ws.column_dimensions[get_column_letter(len(row))].width = 50
        elif field['type'] == constants.DOCUMENT:
            ws.column_dimensions[get_column_letter(len(row))].width = 50
    ws.append(row)
    row_number = 1
    wrap_alignment = openpyxl.styles.Alignment(wrapText=True)
    for proposal in proposals:
        for reviewer in call['reviewers']:
            review = reviews_lookup.get("{} {}".format(proposal['identifier'],
                                                       reviewer))
            if review:
                row_number += 1
                row = [proposal['identifier'], 
                       proposal.get('title') or '',
                       proposal['user'],
                       flask.url_for('review.display',
                                     iuid=review['_id'], _external=True),
                       review.get('finalized') and 'yes' or 'no',
                       reviewer]
                wraptext = []
                hyperlink = []
                for field in call['review']:
                    value = review['values'].get(field['identifier'])
                    if value is None:
                        row.append('')
                    elif field['type'] == constants.TEXT:
                        row.append(value)
                        col = get_column_letter(len(row))
                        wraptext.append(f"{col}{row_number}")
                    elif field['type'] == constants.DOCUMENT:
                        row.append(flask.url_for('review.document',
                                                 iuid=review['_id'],
                                                 document=value,
                                                 _external=True))
                        col = get_column_letter(len(row))
                        hyperlink.append(f"{col}{row_number}")
                    else:
                        row.append(value)
                ws.append(row)
                if wraptext:
                    ws.row_dimensions[row_number].height = 40
                while wraptext:
                    ws[wraptext.pop()].alignment = wrap_alignment
                while hyperlink:
                    colrow = hyperlink.pop()
                    ws[colrow].hyperlink = ws[colrow].value
                    ws[colrow].style = 'Hyperlink'
                colrow = f"{review_column}{row_number}"
                ws[colrow].hyperlink = ws[colrow].value
                ws[colrow].style = 'Hyperlink'
                colrow = f"A{row_number}"
                ws[colrow].hyperlink = flask.url_for('proposal.display',
                                                     pid=ws[colrow].value,
                                                     _external=True)
                ws[colrow].style = 'Hyperlink'
    with tempfile.NamedTemporaryFile(suffix='.xlsx') as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        content = tmp.read()
    return content
