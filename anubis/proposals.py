"Lists of proposals."

import statistics
import tempfile

import flask
import openpyxl
import openpyxl.styles

import anubis.call
import anubis.decision
import anubis.proposal
import anubis.user

from . import constants
from . import utils


blueprint = flask.Blueprint('proposals', __name__)

@blueprint.route('/call/<cid>')
@utils.login_required
def call(cid):
    "List all proposals in a call."
    call = anubis.call.get_call(cid)
    if not call:
        utils.flash_error('No such call.')
        return flask.redirect(flask.url_for('home'))
    if not anubis.call.allow_view(call):
        utils.flash_error("You may not view the call.")
        return flask.redirect(flask.url_for('home'))
    proposals = get_call_proposals(call)
    proposal_bannerfields = [f for f in call['proposal'] if f.get('banner')]
    score_field_ids = [f['identifier'] for f in call['review'] 
                       if f.get('banner') and 
                       f['type'] in constants.NUMERICAL_FIELD_TYPES]
    for proposal in proposals:
        reviews = [r.doc for r in flask.g.db.view('reviews', 'proposal',
                                                  key=proposal['identifier'],
                                                  reduce=False,
                                                  include_docs=True)]
        scores = dict([(id, list()) for id in score_field_ids])
        for review in reviews:
            for id in score_field_ids:
                value = review['values'].get(id)
                if value is not None: scores[id].append(float(value))
        proposal['scores'] = dict()
        for id in score_field_ids:
            proposal['scores'][id] = d = dict()
            try:
                d['mean'] = statistics.mean(scores[id])
            except statistics.StatisticsError:
                d['mean'] = None
            try:
                d['stdev'] = statistics.stdev(scores[id])
            except statistics.StatisticsError:
                d['stdev'] = None
    allow_view_reviews = anubis.call.allow_view_reviews(call)
    allow_view_decisions = anubis.call.allow_view_decisions(call)
    decision_bannerfields = [f for f in call['decision'] if f.get('banner')]
    return flask.render_template('proposals/call.html', 
                                 call=call,
                                 proposals=proposals,
                                 proposal_bannerfields=proposal_bannerfields,
                                 score_field_ids=score_field_ids,
                                 allow_view_reviews=allow_view_reviews,
                                 allow_view_decisions=allow_view_decisions,
                                 decision_bannerfields=decision_bannerfields)

@blueprint.route('/call/<cid>.xlsx')
@utils.login_required
def call_xlsx(cid):
    "Produce an XLSX file of all proposals in a call."
    from openpyxl.utils.cell import get_column_letter
    call = anubis.call.get_call(cid)
    if not call:
        utils.flash_error('No such call.')
        return flask.redirect(flask.url_for('home'))
    if not anubis.call.allow_view(call):
        utils.flash_error("You may not view the call.")
        return flask.redirect(flask.url_for('home'))
    proposals = get_call_proposals(call)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Proposals in call {cid}"
    row = ['Proposal', 'Proposal title', 'Submitter']
    ws.column_dimensions[get_column_letter(2)].width = 30
    for field in call['proposal']:
        row.append(field['identifier'])
        if field['type'] == constants.LINE:
            ws.column_dimensions[get_column_letter(len(row))].width = 40
        elif field['type'] == constants.TEXT:
            ws.column_dimensions[get_column_letter(len(row))].width = 50
        elif field['type'] == constants.DOCUMENT:
            ws.column_dimensions[get_column_letter(len(row))].width = 50
    ws.append(row)
    row_number = 1
    wrap_alignment = openpyxl.styles.Alignment(wrapText=True)
    for proposal in proposals:
        row_number += 1
        row = [proposal['identifier'], 
               proposal.get('title') or '',
               proposal['user']]
        wraptext = []
        hyperlink = []
        for field in call['proposal']:
            value = proposal['values'].get(field['identifier'])
            if value is None:
                row.append('')
            elif field['type'] == constants.TEXT:
                row.append(value)
                col = get_column_letter(len(row))
                wraptext.append(f"{col}{row_number}")
            elif field['type'] == constants.DOCUMENT:
                row.append(flask.url_for('review.document',
                                         iuid=review['_id'],
                                         document=value,
                                         _external=True))
                col = get_column_letter(len(row))
                hyperlink.append(f"{col}{row_number}")
            else:
                row.append(value)
        ws.append(row)
        if wraptext:
            ws.row_dimensions[row_number].height = 40
        while wraptext:
            ws[wraptext.pop()].alignment = wrap_alignment
        while hyperlink:
            colrow = hyperlink.pop()
            ws[colrow].hyperlink = ws[colrow].value
            ws[colrow].style = 'Hyperlink'
        colrow = f"A{row_number}"
        ws[colrow].hyperlink = flask.url_for('proposal.display',
                                             pid=ws[colrow].value,
                                             _external=True)
        ws[colrow].style = 'Hyperlink'
    with tempfile.NamedTemporaryFile(suffix='.xlsx') as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        content = tmp.read()
    response = flask.make_response(content)
    response.headers.set('Content-Type', constants.XLSX_MIMETYPE)
    response.headers.set('Content-Disposition', 'attachment', 
                         filename=f"{cid}_proposals.xlsx")
    return response

@blueprint.route('/user/<username>')
@utils.login_required
def user(username):
    "List all proposals for a user."
    user = anubis.user.get_user(username=username)
    if user is None:
        utils.flash_error('No such user.')
        return flask.redirect(flask.url_for('home'))
    if not anubis.user.allow_view(user):
        utils.flash_error("You may not view the user's proposals.")
        return flask.redirect(flask.url_for('home'))
    return flask.render_template('proposals/user.html',  
                                 user=user,
                                 proposals=get_user_proposals(user['username']),
                                 allow_view_decision=anubis.decision.allow_view)

def get_call_proposals(call):
    "Get the proposals in the call. Only include those allowed to view."
    result = [i.doc for i in flask.g.db.view('proposals', 'call',
                                             key=call['identifier'],
                                             reduce=False,
                                             include_docs=True)]
    return [p for p in result if anubis.proposal.allow_view(p)]

def get_user_proposals(username, call=None):
    "Get all proposals created by the user."
    return [i.doc for i in flask.g.db.view('proposals', 'user',
                                           key=username,
                                           reduce=False,
                                           include_docs=True)]
